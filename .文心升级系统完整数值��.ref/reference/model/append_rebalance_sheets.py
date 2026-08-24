import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
ROOT=Path(r'C:\Users\77522\WorkBuddy\2026-08-01-00-57-25')
X=ROOT/'文心升级系统完整数值表_v2.xlsx'
NPC=ROOT/'feihuaqi-playable'/'config'/'npcs.json'
wb=load_workbook(X)
for n in ['NPC梯度','平衡验收']:
    if n in wb.sheetnames: del wb[n]
BLUE='4472C4'; WHITE='FFFFFF'; THIN=Side(style='thin',color='B7B7B7')
def head(c): c.fill=PatternFill('solid',fgColor=BLUE); c.font=Font(color=WHITE,bold=True); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=Border(bottom=THIN)
def table(ws,ref,name):
    t=Table(displayName=name,ref=ref); t.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True); ws.add_table(t)
# 当前NPC；旧总预算按审计基线固定
old_total={'tongsheng':28,'xiucai':None,'juren':86,'jinshi':105,'zhukaoguan':127}
tiers=json.loads(NPC.read_text(encoding='utf-8'))
ws=wb.create_sheet('NPC梯度')
headers=['档位','NPC ID','姓名','偏科','旧总预算','新总预算','增幅','诗','词','联','笔','学','思','关键能力调整','档位定位']
for i,h in enumerate(headers,1):ws.cell(1,i,h);head(ws.cell(1,i))
rows=[]
for t in tiers:
    for n in t.get('npcs',[]):
        new=sum(n['attrs'].values()); old=old_total[t['id']]
        if t['id']=='xiucai': old=new
        sig=n.get('mech',{}).get('signature',{}); sig=sig.get('main',sig)
        adj='—'
        if n['id']=='ouyang_han': adj='文债耗神 2→3'
        elif n['id'] in ('si_ma_wen','shang_guan_ming','mu_rong_yu'): adj='招牌 +10%→+11%'
        elif n['id']=='xia_hou_jin': adj='稳卷 floorPct 5%→6%'
        elif n['id']=='yuwen_yuan': adj='思力贡献 10%→12%'
        elif n['id']=='wang_shilang': adj='适应阻尼25%→28%；最低收益50%→45%'
        elif n['id']=='li_xue_shi': adj='招牌 +10%→+12%'
        elif n['id']=='zhao_da_ru': adj='稳卷 floorPct 5%→7%'
        a=n['attrs']; rows.append([t['tier'],n['id'],n['name'],n.get('style',''),old,new,(new/old-1) if old else 0,a['shi'],a['ci'],a['lian'],a['bi'],a['xue'],a['si'],adj,t.get('desc','')])
for r,row in enumerate(rows,2):
    for c,v in enumerate(row,1): ws.cell(r,c,v); ws.cell(r,c).alignment=Alignment(horizontal='center' if c not in (14,15) else 'left',vertical='top',wrap_text=c in (14,15))
    ws.cell(r,7).number_format='0.0%'
for c,w in {'N':34,'O':60}.items():ws.column_dimensions[c].width=w
ws.freeze_panes='D2'; ws.auto_filter.ref=f'A1:O{1+len(rows)}';table(ws,f'A1:O{1+len(rows)}','NpcGradientTable');ws.sheet_view.showGridLines=False
# 档位图表辅助区
summary=[]
for t in tiers:
    vals=[sum(n['attrs'].values()) for n in t.get('npcs',[])]; summary.append([t['tier'],sum(vals)/len(vals)])
ws['Q20']='档位';ws['R20']='平均总预算'
for i,(tier,val) in enumerate(summary,21):ws.cell(i,17,tier);ws.cell(i,18,val)
chart=BarChart();chart.title='NPC档位平均总预算';chart.y_axis.title='六维总和';chart.add_data(Reference(ws,min_col=18,min_row=20,max_row=25),titles_from_data=True);chart.set_categories(Reference(ws,min_col=17,min_row=21,max_row=25));chart.height=8;chart.width=14;ws.add_chart(chart,'Q2')
# 验收表
ws=wb.create_sheet('平衡验收')
headers=['类别','指标','调整前/基线','调整后','目标/边界','结果','说明']
for i,h in enumerate(headers,1):ws.cell(1,i,h);head(ws.cell(1,i))
checks=[
['灵感','初始灵感','28','32','适度提高，不覆盖一次抽取+追加骰','PASS','+4，早期可承担一次普通/稀有首级升级'],
['灵感','基础上限','48','54','高于初始且单次最高升级31可支付','PASS','+6；扩容后60或64，二者互斥'],
['灵感','恢复型文心上限','无','T030最多+4；T031最多+6','每枚有局内次数上限','PASS','次数写入v3存档，替换/再获得不刷新'],
['灵感','扩容叠加','无','T032 +6 或 T033 +10','互斥且只结算一次','PASS','替换后不回退；另一枚永久退池'],
['NPC','档位总预算','28/49/86/105/127','28/49/90/117/148','低档不变，进士以上梯度清晰','PASS','约+0%/+0%/+5%/+11%/+17%'],
['仿真','标准玩家封笔率','—','13.6%','不显著高于20%','PASS','sim_feihuaqi N=2000'],
['仿真','标准玩家胜率','—','67.4%','保持可通关且高阶有压力','PASS','sim_feihuaqi N=2000'],
['仿真','新手封笔率','—','16.8%','不显著高于25%','PASS','答题率0.55 N=2000'],
['仿真','新手胜率','—','65.3%','不因低级NPC异常而崩塌','PASS','答题率0.55 N=2000'],
['测试','专项验收','无','39/39','全部通过','PASS','灵感/文心/NPC/存档/UI文案'],
['测试','全量回归','19 suites','21/21 suites','零失败','PASS','含全部sim + save兼容'],
['测试','编辑器冒烟','32','32/32','零失败','PASS','新effect类型与种子同步后'],
['同步','游戏配置=编辑器种子','—','talents true / npcs true','全等','PASS','41枚文心、5档NPC'],
['线上','云端工程覆盖','未知','GitHub Raw 404/不可读','部署前必须确认/同步','BLOCKED','本轮未部署，避免云端旧工程覆盖本地新值']]
for r,row in enumerate(checks,2):
    for c,v in enumerate(row,1):ws.cell(r,c,v);ws.cell(r,c).alignment=Alignment(horizontal='center' if c<7 else 'left',vertical='top',wrap_text=c==7)
    color='C6EFCE' if row[5]=='PASS' else 'FFEB9C';ws.cell(r,6).fill=PatternFill('solid',fgColor=color);ws.cell(r,6).font=Font(bold=True)
ws.column_dimensions['G'].width=60;ws.freeze_panes='A2';ws.auto_filter.ref=f'A1:G{1+len(checks)}';table(ws,f'A1:G{1+len(checks)}','BalanceAcceptanceTable');ws.sheet_view.showGridLines=False
wb.save(X)
print(X);print({s:wb[s].max_row for s in wb.sheetnames})
