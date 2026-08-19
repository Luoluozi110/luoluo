from openpyxl import load_workbook
wb = load_workbook("文心升级系统完整数值表_v2.xlsx", data_only=True)

print("########## 文心总表 (id,name,type,quality,maxLevel,effectType,lv1,max,maxCost,primary) ##########")
ws = wb["文心总表"]
for row in ws.iter_rows(values_only=True):
    if row[0] in (None,"ID"): continue
    print(row[0],"|",row[1],"|",row[2],"|",row[4],"|",row[5],"|",row[7],"|",row[8],"|",row[9],"|",row[10],"|",row[11])

print("\n########## 效果字段映射 ##########")
ws = wb["效果字段映射"]
for row in ws.iter_rows(values_only=True):
    if row[0] is None: continue
    print(row)

print("\n########## 品质成本曲线 (full) ##########")
ws = wb["品质成本曲线"]
for row in ws.iter_rows(values_only=True):
    if row[0] is None: continue
    print(row)
