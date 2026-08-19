import sys, json
try:
    from openpyxl import load_workbook
except Exception as e:
    print("NO_OPENPYXL", e); sys.exit(2)
wb = load_workbook("文心升级系统完整数值表_v2.xlsx", data_only=True)
print("SHEETS:", wb.sheetnames)
for ws in wb.worksheets:
    print("\n===== SHEET:", ws.title, "dims:", ws.dimensions, "=====")
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[:5]:
        print(r)
