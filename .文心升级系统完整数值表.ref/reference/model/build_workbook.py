import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(r"C:\Users\77522\WorkBuddy\2026-08-01-00-57-25")
DATA = ROOT / ".文心升级系统完整数值表.ref" / "reference" / "model" / "data.json"
XLSX = ROOT / "文心升级系统完整数值表_v2.xlsx"

data = json.loads(DATA.read_text(encoding="utf-8"))
wb = Workbook()
wb.remove(wb.active)
for name in ["设计总览","品质成本曲线","文心总表","逐级数值","效果字段映射","实装验收"]:
    wb.create_sheet(name)

BLUE = "4472C4"
WHITE = "FFFFFF"
LIGHT_BLUE = "D9E2F3"
LIGHT_RED = "FCE4D6"
LIGHT_YELLOW = "FFF2CC"
LIGHT_GREEN = "E2F0D9"
ALT = "F7F9FC"
QUALITY = {"普通":"E7E6E6","稀有":"D9EAF7","史诗":"E4DFEC","传说":"FCE4D6"}
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(bottom=THIN)


def head(cell):
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = Font(color=WHITE, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def body(cell, wrap=False, center=False):
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="top", wrap_text=wrap)


def add_table(ws, ref, name):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)


def set_quality_fill(ws, col_letter, start_row, end_row):
    for r in range(start_row, end_row + 1):
        q = ws[f"{col_letter}{r}"].value
        if q in QUALITY:
            fill = PatternFill("solid", fgColor=QUALITY[q])
            ws[f"{col_letter}{r}"].fill = fill

# 设计总览
ws = wb["设计总览"]
ws.merge_cells("A1:E1")
ws["A1"] = data["meta"]["title"]
ws["A1"].font = Font(size=18, bold=True, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor="2F5597")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30
ws.merge_cells("A2:E2")
ws["A2"] = f"生成时间：{data['meta']['generatedAt']}　｜　所有数值均为首轮仿真起点 [PLACEHOLDER]，需接入后进行 Monte Carlo 与实机验证"
ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)
headers = ["模块","项目","内容","数值","单位"]
for c,h in enumerate(headers,1): ws.cell(4,c,h); head(ws.cell(4,c))
rows=[]
rows.append(["Fun Hypothesis","核心乐趣",data["meta"]["funHypothesis"],None,""])
for i,x in enumerate(data["meta"]["assumptions"],1): rows.append(["核心假设",f"假设{i}",x,None,""])
for i,x in enumerate(data["meta"]["designPillars"],1): rows.append(["设计支柱",f"支柱{i}",x,None,""])
for i,x in enumerate(data["meta"]["failureSignals"],1): rows.append(["失败信号",f"信号{i}",x,None,""])
seen=set()
for x in data["qualityCosts"]:
    if x["quality"] in seen: continue
    seen.add(x["quality"])
    rows.append(["品质摘要",x["quality"],f"最大等级 {x['maxLevel']}；目标抽取概率 {x['targetDrawOdds']:.0%}；满级总投入 {x['fullUpgradeCost']} 灵感",x["fullUpgradeCost"],"灵感"])
for r_idx,row in enumerate(rows,5):
    for c_idx,v in enumerate(row,1):
        cell=ws.cell(r_idx,c_idx,v); body(cell,wrap=(c_idx==3),center=(c_idx in (1,2,4,5)))
    if row[0]=="Fun Hypothesis":
        for c in range(1,6): ws.cell(r_idx,c).fill=PatternFill("solid",fgColor=LIGHT_BLUE)
    if row[0]=="失败信号":
        for c in range(1,6): ws.cell(r_idx,c).fill=PatternFill("solid",fgColor=LIGHT_RED)
    if row[0]=="品质摘要":
        for c in range(1,6): ws.cell(r_idx,c).fill=PatternFill("solid",fgColor=QUALITY[row[1]])
ws.column_dimensions["A"].width=16; ws.column_dimensions["B"].width=16; ws.column_dimensions["C"].width=58; ws.column_dimensions["D"].width=12; ws.column_dimensions["E"].width=10
ws.freeze_panes="A5"; ws.auto_filter.ref=f"A4:E{4+len(rows)}"

# 品质成本曲线
ws=wb["品质成本曲线"]
headers=["品质","最大等级","目标抽取概率","当前等级","升至本级成本","升下一级成本","累计升级成本","满级总成本","设计依据"]
for c,h in enumerate(headers,1): ws.cell(1,c,h); head(ws.cell(1,c))
for r_idx,x in enumerate(data["qualityCosts"],2):
    vals=[x["quality"],x["maxLevel"],x["targetDrawOdds"],x["level"],x["costToThis"],x["costToNext"],x["cumulativeCost"],x["fullUpgradeCost"],x["rationale"]]
    for c_idx,v in enumerate(vals,1):
        cell=ws.cell(r_idx,c_idx,v); body(cell,wrap=(c_idx==9),center=(c_idx!=9))
    ws.cell(r_idx,3).number_format="0%"
set_quality_fill(ws,"A",2,1+len(data["qualityCosts"]))
ws.column_dimensions["I"].width=45
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:I{1+len(data['qualityCosts'])}"
add_table(ws,f"A1:I{1+len(data['qualityCosts'])}","QualityCostTable")
# 图表辅助数据
qualities=["普通","稀有","史诗","传说"]
ws["K20"]="等级"
for j,q in enumerate(qualities,12): ws.cell(20,j,q)
for level in range(1,7):
    ws.cell(20+level,11,level)
    for j,q in enumerate(qualities,12):
        rec=next((x for x in data["qualityCosts"] if x["quality"]==q and x["level"]==level),None)
        ws.cell(20+level,j,rec["cumulativeCost"] if rec else None)
chart=LineChart(); chart.title="品质累计升级成本"; chart.y_axis.title="累计灵感"; chart.x_axis.title="等级"; chart.height=8; chart.width=15
chart.add_data(Reference(ws,min_col=12,max_col=15,min_row=20,max_row=26),titles_from_data=True)
chart.set_categories(Reference(ws,min_col=11,min_row=21,max_row=26))
chart.legend.position="r"; ws.add_chart(chart,"K2")
for col in range(11,16): ws.column_dimensions[chr(64+col)].hidden=True

# 文心总表
ws=wb["文心总表"]
headers=["ID","名称","类型","流派","品质","等级上限","旧版等效等级","现有效果类型","Lv1效果","满级效果","满级升级成本","主参数","次参数","当前配置JSON","实装备注"]
for c,h in enumerate(headers,1): ws.cell(1,c,h); head(ws.cell(1,c))
keys=["id","name","kind","school","quality","maxLevel","legacyEquivalentLevel","currentEffectType","level1Effect","maxEffect","fullUpgradeCost","primaryParam","secondaryParam","legacyConfig","implementationNote"]
for r_idx,x in enumerate(data["talentSummary"],2):
    for c_idx,k in enumerate(keys,1):
        cell=ws.cell(r_idx,c_idx,x.get(k,"")); body(cell,wrap=(c_idx in (9,10,14,15)),center=(c_idx in (1,3,4,5,6,7,8,11)))
    if x.get("implementationNote"):
        ws.cell(r_idx,15).fill=PatternFill("solid",fgColor=LIGHT_YELLOW)
    if x["kind"]=="主动": ws.cell(r_idx,3).fill=PatternFill("solid",fgColor="DDEBF7")
set_quality_fill(ws,"E",2,1+len(data["talentSummary"]))
for col,width in {"I":36,"J":36,"N":46,"O":42}.items(): ws.column_dimensions[col].width=width
ws.freeze_panes="C2"; ws.auto_filter.ref=f"A1:O{1+len(data['talentSummary'])}"
add_table(ws,f"A1:O{1+len(data['talentSummary'])}","TalentSummaryTable")

# 逐级数值
ws=wb["逐级数值"]
headers=["ID","名称","类型","品质","等级","等级上限","升至本级成本","升下一级成本","累计升级成本","主参数名","主参数原值","主参数展示","次参数名","次参数原值","次参数展示","使用消耗","完整效果文案","旧版等效级","实装备注"]
for c,h in enumerate(headers,1): ws.cell(1,c,h); head(ws.cell(1,c))
keys=["id","name","kind","quality","level","maxLevel","costToThis","costToNext","cumulativeCost","primaryParam","primaryRaw","primaryDisplay","secondaryParam","secondaryRaw","secondaryDisplay","useCost","effectText","legacyLevel","note"]
for r_idx,x in enumerate(data["levelDetails"],2):
    for c_idx,k in enumerate(keys,1):
        cell=ws.cell(r_idx,c_idx,x.get(k,"")); body(cell,wrap=(c_idx in (17,19)),center=(c_idx not in (10,12,13,15,17,19)))
        if c_idx in (11,14) and isinstance(cell.value,(int,float)): cell.number_format="0.000"
    if x.get("legacyLevel")=="是":
        for c in range(1,20): ws.cell(r_idx,c).fill=PatternFill("solid",fgColor=LIGHT_GREEN)
    else:
        ws.cell(r_idx,4).fill=PatternFill("solid",fgColor=QUALITY[x["quality"]])
for col,width in {"Q":45,"S":45}.items(): ws.column_dimensions[col].width=width
ws.freeze_panes="C2"; ws.auto_filter.ref=f"A1:S{1+len(data['levelDetails'])}"
add_table(ws,f"A1:S{1+len(data['levelDetails'])}","TalentLevelTable")

# 效果字段映射
ws=wb["效果字段映射"]
headers=["effect.type","设计作用","升级字段","单位","建议范围","引擎边界/备注"]
for c,h in enumerate(headers,1): ws.cell(1,c,h); head(ws.cell(1,c))
keys=["type","purpose","fields","unit","range","boundary"]
for r_idx,x in enumerate(data["effectMapping"],2):
    for c_idx,k in enumerate(keys,1):
        cell=ws.cell(r_idx,c_idx,x[k]); body(cell,wrap=(c_idx==6),center=(c_idx in (1,3,4,5)))
    if "P0" in x["boundary"] or "未接线" in x["boundary"]:
        ws.cell(r_idx,6).fill=PatternFill("solid",fgColor="FFC7CE"); ws.cell(r_idx,6).font=Font(color="9C0006",bold=True)
ws.column_dimensions["F"].width=60
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:F{1+len(data['effectMapping'])}"
add_table(ws,f"A1:F{1+len(data['effectMapping'])}","EffectMappingTable")

# 实装验收
ws=wb["实装验收"]
headers=["优先级","模块","实装事项","验收标准"]
for c,h in enumerate(headers,1): ws.cell(1,c,h); head(ws.cell(1,c))
keys=["priority","module","item","acceptance"]
prio_fill={"P0":"FFC7CE","P1":"FFEB9C","P2":"D9EAF7"}
for r_idx,x in enumerate(data["implementationChecklist"],2):
    for c_idx,k in enumerate(keys,1):
        cell=ws.cell(r_idx,c_idx,x[k]); body(cell,wrap=(c_idx in (3,4)),center=(c_idx in (1,2)))
    ws.cell(r_idx,1).fill=PatternFill("solid",fgColor=prio_fill[x["priority"]]); ws.cell(r_idx,1).font=Font(bold=True)
ws.column_dimensions["C"].width=62; ws.column_dimensions["D"].width=58
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:D{1+len(data['implementationChecklist'])}"
add_table(ws,f"A1:D{1+len(data['implementationChecklist'])}","ImplementationTable")

# 全局可读性
for ws in wb.worksheets:
    ws.sheet_view.showGridLines=False
    for row in ws.iter_rows():
        for cell in row:
            if cell.row > 1 and cell.fill.fill_type is None and cell.row % 2 == 0:
                cell.fill=PatternFill("solid",fgColor=ALT)
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.page_setup.orientation="landscape"
    ws.page_setup.fitToWidth=1
    ws.page_setup.fitToHeight=0

wb.save(XLSX)
print(XLSX)
print({ws.title: ws.max_row for ws in wb.worksheets})
