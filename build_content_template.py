# -*- coding: utf-8 -*-
"""生成《飞花棋·内容创意填报表.xlsx》——作者填想法、发回即可快速实现的模板。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:/Users/77522/WorkBuddy/2026-08-01-00-57-25/飞花棋-内容创意填报表.xlsx"

# ---- 样式 ----
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
EX_FILL = PatternFill("solid", fgColor="EEF3F8")      # 示例行
NEW_FILL = PatternFill("solid", fgColor="FFFFFF")      # 新增行（白底）
NOTE_FONT = Font(italic=True, color="666666", size=10)
TITLE_FONT = Font(bold=True, size=15, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def put_table(ws, start_row, headers, rows, widths, example_rows=0):
    """写表头 + 若干行；example_rows 指定前几行为示例（灰底）。"""
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, start_row, len(headers))
    for i, row in enumerate(rows):
        r = start_row + 1 + i
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.alignment = WRAP
            cell.border = BORDER
            cell.fill = EX_FILL if i < example_rows else NEW_FILL
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

# =====================================================================
# Sheet 1 使用说明
# =====================================================================
ws = wb.active
ws.title = "使用说明"
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 110
ws["B1"] = "飞花棋 · 内容创意填报表"
ws["B1"].font = TITLE_FONT
lines = [
    ("", ""),
    ("怎么用", "1) 下面每个 sheet 对应一个内容域（题库 / 奇遇 / 文心 / 叙事文案）。"),
    ("", "2) 灰色「示例」行是格式示范，可整行删掉；在示例下方空白行填写你的新想法。"),
    ("", "3) 文案类字段直接写最终文字；数值/机制类字段按示例格式填（如 shi:3;bi:2）。"),
    ("", "4) 拿不准的字段留空，并在「备注」写你的意图，我会按需补全。"),
    ("", "5) 把填好的文件发回给我，我会直接解析并接入游戏 / 编辑器，无需你再解释结构。"),
    ("", ""),
    ("颜色约定", "蓝底白字=表头；灰底=示例行；白底=你新增的行。"),
    ("重要边界", "本表只收「内容 / 文案 / 数值」创意。若要改玩法规则或引擎逻辑，请在备注写明，我会单独评估工作量。"),
    ("", "流派/段位/弹窗的【数值与机制字段】不可在此改（如 schoolMechanics、评分公式、属性系数），只改其【文案】字段。"),
    ("", ""),
    ("属性速查", "shi=诗力  ci=词力  lian=联力  bi=笔力  xue=学力  si=思力"),
    ("枚举·题库", "type: knowledge(单选)/choice(多向选择)   category: shi/ci/lian/mix   difficulty: 1~3   enabled: true/false"),
    ("枚举·奇遇", "rarity: common/rare/legend   kind: direct(直接给)/challenge(连战胜)/choice(抉择)"),
    ("枚举·文心", "kind: passive(被动)/active(主动)   school: 流派id(通用留空)   effect.type 见文心 sheet 顶部清单"),
    ("", ""),
    ("解析/文案习惯", "叙事用第二人称「你」指称玩家；游戏内会自动把「你」替换成名号。不要自己写 {name} 占位符。"),
    ("", ""),
    ("发回后我能做", "新增题目/奇遇/文心、改文案、调数值、导出对应 json、部署到 GitHub Pages 与 CloudStudio。"),
]
r = 3
for k, v in lines:
    ws.cell(row=r, column=1, value=k).font = Font(bold=True, color="1F4E78")
    c = ws.cell(row=r, column=2, value=v)
    c.alignment = WRAP
    if k in ("重要边界",):
        c.font = Font(color="B00020")
    r += 1

# =====================================================================
# Sheet 2 题库
# =====================================================================
ws = wb.create_sheet("题库")
headers = ["题目ID", "type", "category", "difficulty", "题干 stem",
           "选项1", "属性1", "选项2", "属性2", "选项3", "属性3", "选项4", "属性4",
           "答案序号(仅knowledge)", "解析/点评 analysis", "enabled", "备注"]
rows = [
    ["Q0102", "choice", "mix", 2, "友人远行赴边关，你以何句相赠？",
     "劝君更尽一杯酒，西出阳关无故人", "shi", "黄沙百战穿金甲，不破楼兰终不还", "shi", "赠他一壶家乡水", "null", "", "",
     "", "无标准答案。A 惜别，B 壮行——送别诗的两种姿态。", "true", "示例·多向选择题"],
    ["Q0001", "knowledge", "shi", 2, "「夜半钟声到客船」的作者是？",
     "张继", "", "杜牧", "", "李白", "", "王维", "",
     "0", "答案是张继，出自《枫桥夜泊》。", "true", "示例·知识单选题，答案序号=0(第一项)"],
]
put_table(ws, 1, headers, rows,
          widths=[10, 9, 9, 9, 34, 22, 7, 22, 7, 18, 7, 14, 7, 14, 40, 8, 22],
          example_rows=2)
ws.cell(row=5, column=1, value="▼ 在下面空白行填写你的新题目（ID 留空我帮你生成，如 QNEW001）：").font = NOTE_FONT

# =====================================================================
# Sheet 3 奇遇
# =====================================================================
ws = wb.create_sheet("奇遇")
headers = ["奇遇ID", "name", "rarity", "kind", "奇遇描述 text",
           "直接奖励·属性", "直接奖励·灵感", "直接奖励·文心",
           "挑战场数", "挑战胜利全得·属性", "挑战胜利全得·文心",
           "抉择选项(choice用)", "备注"]
rows = [
    ["E001", "梦笔生花", "legend", "direct", "夜宿江畔驿馆，梦中所执之笔头上忽然生出花来……",
     "bi:5", "", "T007", "", "", "", "", "示例·直接给属性+文心"],
    ["E003", "兰亭修禊", "legend", "challenge", "暮春之初，会于会稽山阴之兰亭。曲水流觞……",
     "", "", "", "3", "bi:5;si:3", "T015", "", "示例·连战3场全胜得奖励"],
    ["E007", "一字之师", "rare", "choice", "你把新作示人，对方沉吟良久，只改了一个字……",
     "", "", "", "", "", "",
     "长揖称谢，从此虚心求教|你依言换去那一字……|xue:3;si:2|+1 || 自恃己见，仍用原句|……|bi:2|-3",
     "示例·抉择：每项=文本|结果文案|属性|灵感，选项间用 || 分隔"],
]
put_table(ws, 1, headers, rows,
          widths=[9, 14, 9, 10, 34, 16, 11, 12, 9, 18, 14, 44, 22],
          example_rows=3)
ws.cell(row=6, column=1, value="▼ 在下面空白行填写新奇遇（kind 决定哪些列要填：direct 填『直接奖励*』；challenge 填『挑战*』；choice 填『抉择选项』）：").font = NOTE_FONT

# =====================================================================
# Sheet 4 文心
# =====================================================================
ws = wb.create_sheet("文心")
headers = ["文心ID", "name", "kind", "school", "文案 text",
           "effect.type", "效果·风格/主属性", "效果·数值 value", "效果·属性包 attrs", "效果·其他参数(JSON)", "备注"]
rows = [
    ["T001", "斗酒诗百篇", "passive", "shixian", "以诗出战获胜时，诗力额外 +1。",
     "on_win_bonus", "shi", "1", "", "", "示例·按文体胜场加属性"],
    ["T004", "博览", "passive", "tongru", "读书破万卷，下笔如有神。学力常驻 +2。",
     "attr_flat", "", "", "xue:2", "", "示例·常驻属性加成"],
    ["T005", "急智", "passive", "qishi", "每场论战的灵感骰点数 +1。",
     "dice_plus", "", "1", "", "", "示例·骰点加成"],
]
put_table(ws, 1, headers, rows,
          widths=[9, 14, 9, 9, 40, 16, 14, 12, 16, 26, 22],
          example_rows=3)
# effect.type 速查清单
ws.cell(row=6, column=1,
        value="▼ effect.type 可选值（直接抄用）：on_win_bonus / attr_flat / dice_plus / dice_mult / crit / "
              "style_pct / theme_pct / start_insp / insp_max / insp_on_win / insp_on_quiz / insp_on_talent / "
              "study_bonus / draw_bonus / comeback / reincarnate / streak_mult / lucky_six / palace_pct / "
              "palace_insp / armory_pct / copy_affinity / fixed_dice / planned_dice / insp_floor / insp_battle_recover").font = NOTE_FONT
ws.cell(row=7, column=1, value="▼ 在下面空白行填写新文心：简单效果只填『风格/数值/属性包』；复杂效果(如 dice_mult 的倍率区间)把参数写进『其他参数(JSON)』列。").font = NOTE_FONT

# =====================================================================
# Sheet 5 叙事文案-流派
# =====================================================================
ws = wb.create_sheet("叙事-流派")
ws.cell(row=1, column=1, value="只改下列【文案】字段（name/motto/flavor/desc）。attr/schoolMechanics/talent/aliases 等数值机制不在此改。").font = NOTE_FONT
headers = ["流派id(不改)", "流派名 name", "口号 motto", "沉浸叙事 flavor(用「你」)", "玩法说明 desc", "备注"]
rows = [
    ["bowen", "博闻", "博观约取，厚积薄发",
     "你自幼好读，藏书万卷皆在腹中。科场之上，你能引百家之言以佐己论……",
     "开局学力 +3，初始文心「博览」。答对考题或完成抉择积累博闻……", "示例"],
    ["qishi", "奇士", "灵台澄澈，万象皆明",
     "你生性爱钻牛角尖，常于无人处反复推敲。奇思往往不循常理……",
     "开局思力 +3，初始文心「推敲」。", "示例"],
]
put_table(ws, 2, headers, rows, widths=[14, 12, 22, 50, 50, 22], example_rows=2)
ws.cell(row=6, column=1, value="▼ 在下面空白行写新流派文案，或改上面示例行的文案；想加全新流派请在备注说明（需引擎支持）。").font = NOTE_FONT

# =====================================================================
# Sheet 6 叙事文案-段位评分
# =====================================================================
ws = wb.create_sheet("叙事-段位评分")
ws.cell(row=1, column=1, value="改【段位/评分的显示文案】。维度评语、段位名/奖励、维度名、加成名+说明、流派分档名+说明、特殊规则说明。").font = NOTE_FONT
headers = ["条目类型", "键 key", "名称(显示名)", "文案(要写/改的文字)", "备注"]
rows = [
    ["维度评语", "wencai", "(维键)", "文采最高：锦心绣口，落笔成章", "comments.<key> 六维最高评语"],
    ["维度评语", "gongli", "(维键)", "功力最高：根柢盘深，厚积薄发", ""],
    ["维度评语", "zhanji", "(维键)", "战绩最高：百战文场，杀伐果断", ""],
    ["维度评语", "qiyu", "(维键)", "奇遇最高：踏遍青山，奇缘满袖", ""],
    ["维度评语", "liupai", "(维键)", "流派最高：一门深入，卓然成家", ""],
    ["维度评语", "yuanman", "(维键)", "圆满最高：从容赴考，功行圆满", ""],
    ["段位档", "tongsheng", "童生", "", "grades[].name；奖励说明写在『文案』列(原 reward 字段)"],
    ["段位档", "wenzong", "文宗", "「文宗」称号与特效", "段位奖励(reward)也在此列填"],
    ["维度名", "wencai", "文采分", "", "dimensions[].name"],
    ["加成名", "sanjuejunheng", "三绝均衡", "诗力、词力、联力均 ≥24", "dimensions[].bonuses[].name + desc"],
    ["流派分档名", "shixian", "诗仙", "诗力 > 词力+联力 且 诗力 ≥30 且 用诗取胜 ≥3 场", "dimensions(liupai).tiers[].name + desc"],
    ["特殊规则说明", "fengbi", "(规则id)", "封笔结局：圆满分记 0，其余五维照常结算", "dimensions(yuanman).specialRules[].desc"],
]
put_table(ws, 2, headers, rows, widths=[14, 16, 14, 52, 40], example_rows=12)
ws.cell(row=16, column=1, value="▼ 在下面空白行：改现有键的文案，或新增维度/加成/分档（新增结构性条目需引擎支持，备注说明即可）。").font = NOTE_FONT

# =====================================================================
# Sheet 7 叙事文案-弹窗
# =====================================================================
ws = wb.create_sheet("叙事-弹窗")
ws.cell(row=1, column=1, value="开局与阶段切换的弹窗文案。『原文案』是当前线上文字，你只需在『新文案』列改写；不改留空。").font = NOTE_FONT
headers = ["弹窗", "字段", "原文案", "新文案(你要改的)", "备注"]
rows = [
    ["prologue(开局序章)", "title", "初入科场", "", ""],
    ["prologue(开局序章)", "text", "你有这么一段模糊的记忆……一鸣惊人。", "", "长文，可整段替换"],
    ["prologue(开局序章)", "button", "踏上征途", "", ""],
    ["zeitgeist(当朝文风)", "title", "风 潮 既 起", "", ""],
    ["zeitgeist(当朝文风)", "lead", "本局科场，文运所钟于二事……", "", ""],
    ["zeitgeist(当朝文风)", "note", "若某场题目恰为热点题材……", "", ""],
    ["zeitgeist(当朝文风)", "button", "谨记于心", "", ""],
    ["stageChange(阶段晋阶)", "names.xiucai", "秀才", "", "阶段名映射，可改显示名"],
    ["stageChange(阶段晋阶)", "names.juren", "举人", "", ""],
    ["stageChange(阶段晋阶)", "names.jinshi", "进士", "", ""],
    ["stageChange(阶段晋阶)", "titleTpl", "{name}阶段 · 晋阶试", "", "{name} 会被替换成阶段名"],
    ["stageChange(阶段晋阶)", "buttonTpl", "进入{name}阶段", "", ""],
    ["stageChange(阶段晋阶)", "default", "基础功名已立……", "", "默认晋阶文案"],
    ["stageChange(阶段晋阶)", "middle", "外圈的试炼已尽……", "", "进中圈文案"],
    ["stageChange(阶段晋阶)", "inner", "中圈的取舍已经定稿……", "", "进内圈文案"],
    ["lap2Intro(会试圈)", "title", "会试圈 · 再入科场", "", ""],
    ["lap2Intro(会试圈)", "text", "童生圈的试炼渐远……", "", ""],
    ["lap2Intro(会试圈)", "button", "进入会试圈", "", ""],
]
put_table(ws, 2, headers, rows, widths=[20, 18, 46, 46, 26], example_rows=18)
ws.cell(row=22, column=1, value="▼ 在『新文案』列直接写替换文字，发回即可生效；也可新增整段弹窗（备注说明触发时机）。").font = NOTE_FONT

wb.save(OUT)
print("已生成:", OUT)
print("sheets:", wb.sheetnames)
